#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
Created on 2012-12-8

@author: Administrator
'''
import sys  
reload(sys)  
sys.setdefaultencoding('utf-8') 

from openpyxl import load_workbook
from struct import pack, unpack
from cStringIO import StringIO
from pyexcel2sc.config import data_sql_dir,data_as_dir, data_dat_dir, excel_dir, isGenerateCreateSql, isGenerateAs,excel_filename
import os,fnmatch

def intPack(value):
    return pack('i', int(value))

def uintPack(value):
    return pack('I', int(value))

def shortPack(value):
    return pack('h', int(value))

def ushortPack(value):
    return pack('H', int(value))

def bytePack(value):
    return pack('b', value)

def ubytePack(value):
    return pack('B', int(value))

def longPack(value):
    return pack('l', long(value))

def ulongPack(value):
    return pack('L', long(value))


def stringPack(*value):
    return pack('H', len(value)) + ''.join(value)

serverClomns = {}
clientClomns = {}

typeDumpDict = {'int':intPack,
            'uint':uintPack,
            'short':shortPack,
            'ushort':ushortPack,
            'byte':bytePack,
            'ubyte':ubytePack,
            'long':longPack,
            'ulong':ulongPack,
            'string':stringPack,
            }
typeDbDict = {'int':'int(10) NOT NULL',
            'uint':'int(10) unsigned NOT NULL',
            'short':'int(10) NOT NULL',
            'ushort':'int(10) unsigned NOT NULL',
            'byte':'int(10) NOT NULL',
            'ubyte':'int(10) unsigned NOT NULL',
            'long':'bigint(20) NOT NULL',
            'ulong':'bigint(20) unsigned NOT NULL',
            'string':'varchar(250) COLLATE utf8_bin NOT NULL',
            }
typeAsReadDict = {'int':'    //%s\n    var %s:int = bytes.readInt();\n',
            'uint':'    //%s\n    var %s:uint = bytes.readUnsignedInt();\n',
            'short':'    //%s\n    var %s:int = bytes.readShort();\n',
            'ushort':'    //%s\n    var %s:uint = bytes.readUnsignedShort();\n',
            'byte':'    //%s\n    var %s:int = bytes.readByte();\n',
            'ubyte':'    //%s\n    var %s:uint = bytes.readUnsignedByte();\n',
            'long':'    //%s\n    var %s:int = bytes.readInt();bytes.readInt();//as not surport long\n',
            'ulong':'    //%s\n    var %s:uint = bytes.readUnsignedInt();bytes.readUnsignedInt();//as not surport long\n',
            'string':'    //%s\n    var length:uint=bytes.readUnsignedShort(); \n    if(length) {\n        var %s:String = bytes.readMultiByte(length,CharCode.UTF8);\n    }\n',
            }

def processdomain(clomn, value):
    value = value.lower()
    if value == 'sc':
        serverClomns[clomn] = None
        clientClomns[clomn] = None
    elif value == 's':
        serverClomns[clomn] = None
    elif value == 'c':
        clientClomns[clomn] = None
        
def processtype(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn] = [value,typeDbDict[value]]
    if clientClomns.has_key(clomn):
        clientClomns[clomn] = [value,typeDumpDict[value]]

def processClomnName(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn].append(value)
    if clientClomns.has_key(clomn):
        clientClomns[clomn].append(value)

def processClomnNameComment(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn].append(value)
    if clientClomns.has_key(clomn):
        clientClomns[clomn].append(value)

def processRow(rowIndex, row, rowDict):
    for clomn, cell in enumerate(row):
            value = cell.internal_value
            if cell.data_type == 'n':
                value = long(value)
            if(rowIndex == 1):
                if not value:
                    print '第一行存在空值'
                    return
                processdomain(clomn, value)
                continue
            if(rowIndex == 2):
                if not value:
                    print '第二行存在空值'
                    return
                processtype(clomn, value)
                continue
            if rowIndex == 3:
                processClomnName(clomn, value)
                continue
            if rowIndex == 4:
                processClomnNameComment(clomn, value)
                continue
            datatype = serverClomns.get(clomn)
            if not datatype:
                datatype = clientClomns.get(clomn)
            if datatype and datatype[0]  == 'string':
                value = str(value)
            rowDict[clomn] = value

def generateCreateSql():
    createSql = 'drop table if exists `m_%s`;\nCREATE TABLE `m_%s` (\n'%(currentFileName,currentFileName)
    for k,v in serverClomns.iteritems():
        if k==0:
            primaryClomn= v[2]
        createSql += '    `%s` %s,\n'%(v[2],v[1])
    createSql +='    PRIMARY KEY (`%s`)\n'%primaryClomn
    createSql += ') ENGINE=MyISAM DEFAULT CHARSET=utf8 COLLATE=utf8_bin;\n\n'
    return createSql

def generateAs():
    asCode = 'var bytes:ByteArray = res.data;\nbytes.endian=Endian.LITTLE_ENDIAN;\nbytes.position=0;\nvar listLength:int=bytes.readUnsignedShort();\nfor(var i:uint=0;i<listLength;i++){\n'
    for k,v in clientClomns.iteritems():
        line = typeAsReadDict[v[0]]
        asCode += line%(v[3],v[2])
    asCode += "}"
    return asCode

def generateInsertHead():
    return 'insert into `m_%s` values ('%(currentFileName)     

def packDataList(dataList):
    serverStr = StringIO()
    clientStr = StringIO()
    asStr = StringIO()
    if isGenerateCreateSql:
        serverStr.write(generateCreateSql())
    if isGenerateAs:
        asStr.write(generateAs())
        f = open(data_as_dir + '/'+currentFileName+'.as', 'w')
        f.write(asStr.getvalue())
        f.close()
    print  len(dataList)
    clientStr.write(pack('H', len(dataList)))
    for rowIndex, row in enumerate(dataList):
        insertHead = generateInsertHead()
        for clomn, cell in row.iteritems():
            serverClomnsInfo = serverClomns.get(clomn)
            if serverClomnsInfo:
                if('string' == serverClomnsInfo[0]):
                    cell = cell.encode('utf-8','ignore')
                    insertHead += '\'' + cell+'\', '
                else:
                    insertHead += str(long(cell)) +', '
            clientClomnsInfo = clientClomns.get(clomn)
            if clientClomnsInfo:
                clientClomnsPack = clientClomnsInfo[1]
                if('string' == clientClomnsInfo[0] and clientClomnsPack):
                    cell = cell.encode('utf-8','ignore')
                    clientStr.write(clientClomnsPack(*cell))
                elif clientClomnsPack:
                    clientStr.write(clientClomnsPack(cell))
            print rowIndex, row
        insertHead = insertHead[0:len(insertHead)-2] + ');\n'
        serverStr.write(insertHead)
    f = open(data_sql_dir + '/m_'+currentFileName+'.sql', 'w')
    f.write(serverStr.getvalue())
    f.close()
    f = open(data_dat_dir + '/'+currentFileName+'.dat', 'wb')
    f.write(clientStr.getvalue())
    f.close()
def parseExcel(excelFile):
    serverClomns.clear()
    clientClomns.clear()
    wb = load_workbook(filename = excelFile, use_iterators = True)
    ws = wb.get_sheet_by_name(name = wb.get_sheet_names()[0]) # ws is now an IterableWorksheet
    rowIndex = 0
    sheetRowList = []
    for row in ws.iter_rows(): # it brings a new method: iter_rows()
        rowIndex += 1
        rowCellDict = {}
        processRow(rowIndex, row,rowCellDict)
        if rowCellDict:
            sheetRowList.append(rowCellDict)
    
    packDataList(sheetRowList)
    
                    
if __name__ == '__main__':
    patterns = ['*.xls', '*.xlsx']
    for root, dirs, files in os.walk(excel_dir, True):
        for name in files:
            for pattern in patterns:
                if fnmatch.fnmatch(name,pattern):
                    print name
                    currentFileName= os.path.splitext(name)[0]
                    excelFile = os.path.join(root,name)
                    if excel_filename == '*' or (excel_filename != '*' and excel_filename == name):
                        parseExcel(excelFile)
                        break
    
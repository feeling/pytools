#!/usr/bin/env python
# -*- coding: utf-8 -*-
'''
Created on 2012-12-8

@author: Administrator
'''
import sys  
from symbol import except_clause
from lib.dir_utils import get_main_dir
reload(sys)  
sys.setdefaultencoding('utf-8') 

from openpyxl import load_workbook
from struct import pack, unpack
from cStringIO import StringIO
from excel2sc.config import data_sql_dir,data_as_dir, data_dat_dir, excel_dir, isGenerateCreateSql, isGenerateAs,excel_filename
import os,fnmatch

def intPack(value):
    return pack('i', int(value))

def uintPack(value):
    return pack('I', int(value))

def shortPack(value):
    return pack('h', int(value))

def ushortPack(value):
    return pack('H', int(value))

def bytePack(value):
    return pack('b', value)

def ubytePack(value):
    return pack('B', int(value))

def longPack(value):
    return pack('l', long(value))

def ulongPack(value):
    return pack('L', long(value))


def stringPack(*value):
    return pack('H', len(value)) + ''.join(value)

serverClomns = {}
clientClomns = {}

typeDumpDict = {'int':intPack,
            'uint':uintPack,
            'short':shortPack,
            'ushort':ushortPack,
            'byte':bytePack,
            'ubyte':ubytePack,
            'long':longPack,
            'ulong':ulongPack,
            'string':stringPack,
            }
typeDbDict = {'int':'int(10) NOT NULL',
            'uint':'int(10) unsigned NOT NULL',
            'short':'int(10) NOT NULL',
            'ushort':'int(10) unsigned NOT NULL',
            'byte':'int(10) NOT NULL',
            'ubyte':'int(10) unsigned NOT NULL',
            'long':'bigint(20) NOT NULL',
            'ulong':'bigint(20) unsigned NOT NULL',
            'string':'varchar(250) COLLATE utf8_bin NOT NULL',
            }
typeAsReadDict = {'int':'    //%s\n    var %s:int = bytes.readInt();\n',
            'uint':'    //%s\n    var %s:uint = bytes.readUnsignedInt();\n',
            'short':'    //%s\n    var %s:int = bytes.readShort();\n',
            'ushort':'    //%s\n    var %s:uint = bytes.readUnsignedShort();\n',
            'byte':'    //%s\n    var %s:int = bytes.readByte();\n',
            'ubyte':'    //%s\n    var %s:uint = bytes.readUnsignedByte();\n',
            'long':'    //%s\n    var %s:Number = bytes.readDouble();\n',
            'ulong':'    //%s\n    var %s:Number = bytes.readDouble();\n',
            'string':'    //%s\n    length = bytes.readUnsignedShort(); \n    if(length) {\n        var %s:String = bytes.readMultiByte(length,CharCode.UTF8);\n    }\n',
            }

all_server_sql = []

def processdomain(clomn, value):
    value = value.lower()
    if value == 'sc':
        serverClomns[clomn] = None
        clientClomns[clomn] = None
    elif value == 's':
        serverClomns[clomn] = None
    elif value == 'c':
        clientClomns[clomn] = None
        
def processtype(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn] = [value,typeDbDict[value]]
    if clientClomns.has_key(clomn):
        clientClomns[clomn] = [value,typeDumpDict[value]]

def processClomnName(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn].append(value)
    if clientClomns.has_key(clomn):
        clientClomns[clomn].append(value)

def processClomnNameComment(clomn, value):
    value = value.lower()
    if serverClomns.has_key(clomn):
        serverClomns[clomn].append(value)
    if clientClomns.has_key(clomn):
        clientClomns[clomn].append(value)

def processRow(rowIndex, row, rowDict):
    for clomn, cell in enumerate(row):
            value = cell.internal_value
            if cell.data_type == 'n':
                value = long(value) if value else value
            if(rowIndex == 1):
                if not value:
                    print '第一行存在空值'
                    return
                processdomain(clomn, value)
                continue
            if(rowIndex == 2):
                if not value:
                    print '第二行存在空值'
                    return
                processtype(clomn, value)
                continue
            if rowIndex == 3:
                processClomnName(clomn, value)
                continue
            if rowIndex == 4:
                processClomnNameComment(clomn, value)
                continue
            datatype = serverClomns.get(clomn)
            if not datatype:
                datatype = clientClomns.get(clomn)
            if datatype and datatype[0]  == 'string':
#                print clomn, value
#                if not value: 
                if (not value) and (str(value) != '0.0') : # gai # 或许还会有问题,先用用看吧,不行了继续改
                    value = ''
                if str(value) == '0.0':
                    value = '0'
                value = str(value) 
            elif datatype and datatype[0] and not value:
                value = 0                
            rowDict[clomn] = value

def generateCreateSql():
    createSql = 'drop table if exists `m_%s`;\nCREATE TABLE `m_%s` (\n'%(currentFileName,currentFileName)
    for k,v in serverClomns.iteritems():
        if k==0:
            primaryClomn= v[2]
        createSql += '    `%s` %s,\n'%(v[2],v[1])
    createSql +='    PRIMARY KEY (`%s`)\n'%primaryClomn
    createSql += ') ENGINE=MyISAM DEFAULT CHARSET=utf8 COLLATE=utf8_bin;\n\n'
    return createSql

def generateAs():
    asCode = 'var length:int = 0;\nvar bytes:ByteArray = res.data;\nbytes.endian=Endian.LITTLE_ENDIAN;\nbytes.position=0;\nvar listLength:int=bytes.readUnsignedShort();\nfor(var i:uint=0;i<listLength;i++){\n'
    for k,v in clientClomns.iteritems():
        line = typeAsReadDict[v[0]]
        asCode += line%(v[3],v[2])
    asCode += "}"
    return asCode

def generateInsertHead():
    return 'insert into `m_%s` values ('%(currentFileName)     

def packDataList(dataList):
    serverStr = StringIO()
    clientStr = StringIO()
    asStr = StringIO()
    if isGenerateCreateSql and serverClomns:
        serverStr.write(generateCreateSql())
    if isGenerateAs and clientClomns:
        asStr.write(generateAs())
        f = open(currentdir + data_as_dir + '/'+currentFileName+'.as', 'w')
        f.write(asStr.getvalue())
        f.close()
    print  len(dataList)
    clientStr.write(pack('H', len(dataList)))
    for rowIndex, row in enumerate(dataList):
        insertHead = generateInsertHead()
        row_data_list = []
        for clomn, cell in row.iteritems():
            serverClomnsInfo = serverClomns.get(clomn)
            if serverClomnsInfo:
                try:
                    if('string' == serverClomnsInfo[0]):
                        cell = cell.encode('utf-8','ignore')
                        insertHead += '\'' + cell+'\', '
                        row_data_list.append(cell)
                    else:
                        insertHead += str(long(cell)) +', '
                        row_data_list.append(str(long(cell)))
                except Exception, e:
                    print '%s行%s列 数据值为：%s，无法解析，错误信息：%s'%(rowIndex, clomn, cell, e)
            clientClomnsInfo = clientClomns.get(clomn)
            if clientClomnsInfo:
                clientClomnsPack = clientClomnsInfo[1]
                try:
                    if 'string' == clientClomnsInfo[0] and clientClomnsPack:
                        cell = cell.encode('utf-8', 'ignore')
                        client_row = clientClomnsPack(*cell)
                    elif clientClomnsPack:
                        client_row = clientClomnsPack(cell)
                except Exception, e:
                    print '%s行%s列 数据值为：%s，无法解析，错误信息：%s'%(rowIndex, clomn, cell, e)
        insertHead = insertHead[0:len(insertHead)-2] + ');\n'
        row_blank = True
        for i in row_data_list:
            if i and '0' != i:
                row_blank = False
        if not row_blank:
            if serverClomnsInfo:
                serverStr.write(insertHead)
            if clientClomnsInfo:
                clientStr.write(client_row)
    if isGenerateCreateSql and serverClomns:
        f = open(currentdir + data_sql_dir + '/m_'+currentFileName+'.sql', 'w')
        f.write(serverStr.getvalue())
        f.close()
        all_server_sql.append(serverStr.getvalue())
    if clientClomns:
        f = open(currentdir + data_dat_dir + '/'+currentFileName+'.dat', 'wb')
        f.write(clientStr.getvalue())
        f.close()
def parseExcel(excelFile):
    try:
        serverClomns.clear()
        clientClomns.clear()
        wb = load_workbook(filename = excelFile, use_iterators = True)
        ws = wb.get_sheet_by_name(name = wb.get_sheet_names()[0]) # ws is now an IterableWorksheet
        rowIndex = 0
        sheetRowList = []
        for row in ws.iter_rows(): # it brings a new method: iter_rows()
            rowIndex += 1
            rowCellDict = {}
            processRow(rowIndex, row,rowCellDict)
            if rowCellDict:
                sheetRowList.append(rowCellDict)
        
        packDataList(sheetRowList)
    except Exception,e:
        print e
    
if __name__ == '__main__':
    global currentdir
    global currentFileName
    try:
        currentdir = get_main_dir()+'/'
        print 'currentdir:', currentdir
        os.chdir(currentdir)
        patterns = ['*.xls', '*.xlsx']
        if excel_filename != '*':
            patterns = excel_filename.split(',')
        for root, dirs, files in os.walk(excel_dir, True):   
            for name in files:
                for pattern in patterns:
                    if fnmatch.fnmatch(name,pattern):
                        currentFileName= os.path.splitext(name)[0]
                        excelFile = os.path.join(root,name)
                        print excelFile
                        parseExcel(excelFile)
                        break
    except Exception,e:
        print
    if isGenerateCreateSql and serverClomns:
        f = open(currentdir + data_sql_dir + '/m_'++'.sql', 'w')
        for i in all_server_sql:
            f.write(i)
        f.close()
    raw_input('input enter...>') 
